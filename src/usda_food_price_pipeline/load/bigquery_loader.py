"""Phase 2: load the Phase-1 raw files into BigQuery raw tables.

Reads the raw files produced by Phase 1 under ``data/raw/`` and batch-loads them,
essentially **as-is**, into three clearly named raw tables:

    raw_nutrition     <- data/raw/nutrition/*.json        (FDC /foods/search pages)
    raw_prices_bls    <- data/raw/prices/bls/*.json        (BLS APU response)
    raw_prices_fmap   <- data/raw/prices/fmap/*.xlsx       (ERS F-MAP workbooks)

Only the *minimal* reshaping needed to make each source loadable is done here:
  - nutrition: pull the ``foods`` array out of each search page (one row per food);
  - BLS:       flatten ``Results.series[].data[]`` into one row per observation;
  - F-MAP:     read each .xlsx sheet's rows (header row -> a JSON record per row).
The original payload is preserved (``raw_json`` columns) — no cleaning, type
casting, standardizing, or joining. That is Phase 3 (dbt).

Idempotency: every table is loaded with ``WRITE_TRUNCATE`` in a single batch load
job per source, so re-running fully replaces the table — no duplicate rows.
Batch loads are free; we never use streaming inserts.

Run (as a module, with the src/ layout on the path):
    # PowerShell:  $env:PYTHONPATH = "src"
    # bash:        export PYTHONPATH=src
    python -m usda_food_price_pipeline.load.bigquery_loader
    python -m usda_food_price_pipeline.load.bigquery_loader --dry-run   # parse only, no BigQuery

Needs ``GOOGLE_APPLICATION_CREDENTIALS`` (service-account JSON) in ``.env``.
The target project is taken from that service account (override with
``BIGQUERY_PROJECT``); dataset defaults to ``usda_raw`` (override with
``BIGQUERY_DATASET`` or ``--dataset``).
"""

from __future__ import annotations

import argparse
import json
import os
from datetime import datetime, timezone
from pathlib import Path

from ..ingestion import common

DEFAULT_DATASET = "usda_raw"
DEFAULT_LOCATION = "US"  # BigQuery free tier (10 GB storage, 1 TB query/mo) lives in US.

TABLE_NUTRITION = "raw_nutrition"
TABLE_BLS = "raw_prices_bls"
TABLE_FMAP = "raw_prices_fmap"


# --------------------------------------------------------------------------- #
# Pure row builders (no BigQuery, no I/O) — unit-tested in tests/.
# --------------------------------------------------------------------------- #
def nutrition_rows_from_page(page: dict, source_file: str, loaded_at: str) -> list[dict]:
    """One row per food in an FDC ``/foods/search`` page.

    Keeps the whole food object verbatim in ``raw_json``; promotes ``fdcId`` to a
    column so the load is easy to verify and key on later (no cleaning).
    """
    rows = []
    for food in page.get("foods", []):
        rows.append(
            {
                "source_file": source_file,
                "fdc_id": food.get("fdcId"),
                "raw_json": json.dumps(food, ensure_ascii=False),
                "loaded_at": loaded_at,
            }
        )
    return rows


def bls_rows_from_response(data: dict, source_file: str, loaded_at: str) -> list[dict]:
    """Flatten ``Results.series[].data[]`` into one row per observation.

    Values are kept as the API returned them (strings); ``footnotes`` (a list) is
    preserved as a JSON string. No casting or cleaning.
    """
    rows = []
    for series in data.get("Results", {}).get("series", []):
        series_id = series.get("seriesID")
        for obs in series.get("data", []):
            rows.append(
                {
                    "source_file": source_file,
                    "series_id": series_id,
                    "year": obs.get("year"),
                    "period": obs.get("period"),
                    "period_name": obs.get("periodName"),
                    "value": obs.get("value"),
                    "latest": obs.get("latest"),
                    "footnotes": json.dumps(obs.get("footnotes", []), ensure_ascii=False),
                    "loaded_at": loaded_at,
                }
            )
    return rows


def fmap_rows_from_sheet(sheet_name, sheet_rows, source_file: str, loaded_at: str) -> list[dict]:
    """Turn one worksheet's rows into raw rows.

    ``sheet_rows`` is an iterable of cell-value tuples (the first is treated as the
    header). Each data row becomes a ``{header: cell}`` dict stored verbatim in
    ``raw_json`` (cells that aren't JSON-native, e.g. dates, fall back to str).
    """
    it = iter(sheet_rows)
    try:
        raw_header = next(it)
    except StopIteration:
        return []  # empty sheet

    header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(raw_header)]
    rows = []
    for idx, values in enumerate(it, start=1):
        record = {
            (header[i] if i < len(header) else f"col_{i}"): values[i]
            for i in range(len(values))
        }
        rows.append(
            {
                "source_file": source_file,
                "sheet_name": str(sheet_name),
                "row_index": idx,
                "raw_json": json.dumps(record, ensure_ascii=False, default=str),
                "loaded_at": loaded_at,
            }
        )
    return rows


# --------------------------------------------------------------------------- #
# File readers (I/O, but still no BigQuery).
# --------------------------------------------------------------------------- #
def _load_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def collect_nutrition_rows(loaded_at: str) -> tuple[list[dict], int]:
    """Read every nutrition page file. Returns (rows, file_count)."""
    files = sorted(common.raw_dir("nutrition").glob("*.json"))
    rows: list[dict] = []
    for path in files:
        rows.extend(nutrition_rows_from_page(_load_json(path), path.name, loaded_at))
    return rows, len(files)


def collect_bls_rows(loaded_at: str) -> tuple[list[dict], int]:
    """Read every BLS response file. Returns (rows, file_count)."""
    files = sorted(common.raw_dir("prices", "bls").glob("*.json"))
    rows: list[dict] = []
    for path in files:
        rows.extend(bls_rows_from_response(_load_json(path), path.name, loaded_at))
    return rows, len(files)


def collect_fmap_rows(loaded_at: str) -> tuple[list[dict], int]:
    """Read every F-MAP workbook (all sheets). Returns (rows, file_count)."""
    import openpyxl

    files = sorted(common.raw_dir("prices", "fmap").glob("*.xlsx"))
    rows: list[dict] = []
    for path in files:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows.extend(
                    fmap_rows_from_sheet(
                        sheet_name, ws.iter_rows(values_only=True), path.name, loaded_at
                    )
                )
        finally:
            wb.close()
    return rows, len(files)


# --------------------------------------------------------------------------- #
# BigQuery helpers (lazy import so the module + pure functions work without the
# client installed, e.g. during unit tests).
# --------------------------------------------------------------------------- #
def _schemas(bq):
    """Explicit schemas (autodetect off) keyed by table name."""
    SF = bq.SchemaField
    return {
        TABLE_NUTRITION: [
            SF("source_file", "STRING"),
            SF("fdc_id", "INT64"),
            SF("raw_json", "STRING"),
            SF("loaded_at", "TIMESTAMP"),
        ],
        TABLE_BLS: [
            SF("source_file", "STRING"),
            SF("series_id", "STRING"),
            SF("year", "STRING"),
            SF("period", "STRING"),
            SF("period_name", "STRING"),
            SF("value", "STRING"),
            SF("latest", "STRING"),
            SF("footnotes", "STRING"),
            SF("loaded_at", "TIMESTAMP"),
        ],
        TABLE_FMAP: [
            SF("source_file", "STRING"),
            SF("sheet_name", "STRING"),
            SF("row_index", "INT64"),
            SF("raw_json", "STRING"),
            SF("loaded_at", "TIMESTAMP"),
        ],
    }


def ensure_dataset(client, bq, dataset_id: str, location: str):
    """Create the dataset if it doesn't already exist (idempotent)."""
    from google.api_core.exceptions import NotFound

    ref = f"{client.project}.{dataset_id}"
    try:
        return client.get_dataset(ref)
    except NotFound:
        dataset = bq.Dataset(ref)
        dataset.location = location
        ds = client.create_dataset(dataset)
        print(f"  created dataset {ref} (location={location})")
        return ds


def load_table(client, bq, table_id: str, schema, rows: list[dict]) -> int:
    """Batch-load ``rows`` into ``table_id`` with WRITE_TRUNCATE. Returns row count.

    WRITE_TRUNCATE replaces the table contents every run, so re-running never
    creates duplicates. Uses a load job (free), not streaming inserts.
    """
    job_config = bq.LoadJobConfig(
        schema=schema,
        write_disposition=bq.WriteDisposition.WRITE_TRUNCATE,
        source_format=bq.SourceFormat.NEWLINE_DELIMITED_JSON,
    )
    job = client.load_table_from_json(rows, table_id, job_config=job_config)
    job.result()  # wait for the batch load to finish
    return client.get_table(table_id).num_rows  # table metadata, not a billed query


# --------------------------------------------------------------------------- #
# Orchestration / CLI.
# --------------------------------------------------------------------------- #
SOURCES = ("nutrition", "bls", "fmap")

_COLLECTORS = {
    "nutrition": (TABLE_NUTRITION, collect_nutrition_rows),
    "bls": (TABLE_BLS, collect_bls_rows),
    "fmap": (TABLE_FMAP, collect_fmap_rows),
}


def run(dataset_id: str, location: str, only=SOURCES, dry_run: bool = False) -> int:
    """Build rows for each source and (unless --dry-run) load them into BigQuery."""
    loaded_at = datetime.now(timezone.utc).isoformat()

    # Read + reshape all requested sources first (cheap, no BigQuery yet).
    built = {}
    for source in only:
        table, collector = _COLLECTORS[source]
        rows, n_files = collector(loaded_at)
        built[source] = (table, rows)
        print(f"  {source:9s}: {n_files} file(s) -> {len(rows):,} row(s) for {table}")

    if dry_run:
        print("\nDry run: parsed files only, nothing written to BigQuery.")
        _print_counts({table: len(rows) for table, rows in built.values()})
        return 0

    from google.cloud import bigquery

    project = os.environ.get("BIGQUERY_PROJECT")  # else inferred from the credentials
    client = bigquery.Client(project=project) if project else bigquery.Client()
    print(f"\nProject: {client.project}  Dataset: {dataset_id}  Location: {location}")
    ensure_dataset(client, bigquery, dataset_id, location)
    schemas = _schemas(bigquery)

    counts = {}
    for table, rows in built.values():
        table_id = f"{client.project}.{dataset_id}.{table}"
        if not rows:
            print(f"  WARNING: no rows for {table}; skipping load (run Phase-1 ingestion first?)")
            continue
        n = load_table(client, bigquery, table_id, schemas[table], rows)
        counts[table] = n
        print(f"  loaded {table_id}: {n:,} rows (WRITE_TRUNCATE)")

    _print_counts(counts)
    return 0


def _print_counts(counts: dict) -> None:
    print("\nRow counts per raw table:")
    for table in (TABLE_NUTRITION, TABLE_BLS, TABLE_FMAP):
        if table in counts:
            print(f"  {table:18s} {counts[table]:>10,}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Load Phase-1 raw files into BigQuery raw tables (idempotent, batch loads)."
    )
    parser.add_argument(
        "--dataset",
        default=os.environ.get("BIGQUERY_DATASET", DEFAULT_DATASET),
        help=f"BigQuery dataset for the raw tables (default: {DEFAULT_DATASET}).",
    )
    parser.add_argument(
        "--location",
        default=os.environ.get("BIGQUERY_LOCATION", DEFAULT_LOCATION),
        help=f"Dataset location if it must be created (default: {DEFAULT_LOCATION}).",
    )
    parser.add_argument(
        "--only",
        nargs="+",
        choices=SOURCES,
        default=list(SOURCES),
        help="Load only these sources (default: all).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse the raw files and print row counts without touching BigQuery.",
    )
    args = parser.parse_args(argv)

    common.load_environment()
    if not args.dry_run and not os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"):
        print(
            "ERROR: GOOGLE_APPLICATION_CREDENTIALS is not set (.env). "
            "See .env.example. Use --dry-run to parse without BigQuery.",
        )
        return 1

    print("Loading raw files into BigQuery" + (" [DRY RUN]" if args.dry_run else ""))
    return run(args.dataset, args.location, only=args.only, dry_run=args.dry_run)


if __name__ == "__main__":
    raise SystemExit(main())
